import json
import csv
from openpyxl import Workbook

# with open("bg_csv.csv", 'w', newline="") as f:  # newline每行写入时的符号
#     writer = csv.writer(f)
#     writer.writerow(["name", "age", "sex", "score"])  # 头部信心
#     data = (
#         ["owen", '21', 'male', '67'],
#         ['lacks', '22', 'female', '89'],
#         ['rose', '20', 'female', '80'],
#         ['harry', '23', 'male', '76'],
#         ['jeyyr', '21', 'male', '65'],
#         ('pawn', '23', 'male', '86')
#     )
#     writer.writerows(data)

# wb = Workbook()  # 新建一个工作簿
# ws = wb.worksheets[0]  # 新建第一个工作表
# ws.title = "file.csv"
# # 在工作簿中填入值
# ws.cell(row=1, column=1, value="name")
# ws.cell(row=1, column=2, value="score")
# ws.cell(row=2, column=1, value="owen")
# ws.cell(row=2, column=2, value="87")
# ws.cell(row=3, column=1, value="messi")
# ws.cell(row=3, column=2, value="97")
# ws.cell(row=4, column=1, value="moon")
# ws.cell(row=4, column=2, value="94")
# wb.save(filename="file.xlxs")  # 存储工作簿

# stu = {"name": "owen", "age": 21}
# print(type(stu))
# json_str = json.dumps(stu)  # 将字典转化为字符串
# print(type(json_str))

