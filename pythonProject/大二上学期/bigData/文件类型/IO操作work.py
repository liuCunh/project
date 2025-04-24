import json
import csv
from openpyxl import Workbook
import time
import random

# 第一题
with open("1.txt", 'w', encoding="utf8") as f:
    f.write("刘春 20215670")

# 第二题
wb = Workbook()
ws = wb.worksheets[0]
ws.title = "刘春"
sex = {1: "male", 2: "female"}
ws.cell(1, 1, "name")
ws.cell(1, 2, "Id")
ws.cell(1, 3, "sex")
for i in range(2, 11):
    for j in range(1, 4):
        if j == 1:
            value = (chr(95+i)) * random.randint(3, 6)
        elif j == 2:
            value = f"10{i}"
        else:
            ran = random.randint(1, 2)
            value = sex[ran]
        ws.cell(i, j, value)
        time.sleep(0.1)
wb.save(filename="file.xlsx")

# 第三题
with open("刘春.csv", "w", newline="") as xls:
    writer = csv.writer(xls)
    writer.writerow(["name", "score"])
    data = [
        ["owen", 89],
        ['lacks', 90],
        ['rose', 65],
        ['harry', 72],
        ['jeyyr', 88],
        ['pawn', 60],
        ["Tom", 92],
        ["jerry", 86],
        ["lily", 87],
        ["lucy", 59]

    ]
    writer.writerows(data)

# json转string
print(type(sex))
sex = json.dumps(sex)
print(type(sex))